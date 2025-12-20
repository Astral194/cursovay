import io
import openpyxl

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from .forms import LoginForm, AddAdminForm, AddDoctorForm
from .decorators import login_required
from django.db import transaction, connection
from .models import (
    SystemUser,
    Doctor,
    Patient,
    Alias,
    Visit,
    MedicalRecord,
    Diagnosis,
    Prescription,
    Medication,
    LabTest,
    ActionLog,
    EncryptionKey
)

from .constants import ADD_ALLOWED_TABLES

TABLES = {
    "system_users": SystemUser,
    "doctors": Doctor,
    "patients": Patient,
    "aliases": Alias,
    "encryption_keys": EncryptionKey,
    "visits": Visit,
    "medical_records": MedicalRecord,
    "diagnoses": Diagnosis,
    "prescriptions": Prescription,
    "medications": Medication,
    "lab_tests": LabTest,
    "action_logs": ActionLog,
}


def authenticate_user(email, plain_password):
    with connection.cursor() as cursor:
        cursor.execute("""
                       SELECT id, role
                       FROM system_users
                       WHERE email = %s
                         AND hashed_password = crypt(%s, hashed_password) LIMIT 1
                       """, [email, plain_password])

        row = cursor.fetchone()

    return row


def login_view(request):
    error = None

    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]
            password = form.cleaned_data["password"]

            result = authenticate_user(email, password)

            if result:
                user_id, role = result
                request.session["user_id"] = user_id
                request.session["user_role"] = role
                return redirect("dashboard")
            else:
                error = "Неверный email или пароль"
    else:
        form = LoginForm()

    return render(request, "login.html", {
        "form": form,
        "error": error
    })


def logout_view(request):
    with connection.cursor() as cursor:
        cursor.execute("RESET ROLE")
    request.session.flush()
    return redirect("login")


@login_required
def dashboard(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login")

    try:
        user = SystemUser.objects.get(id=user_id)
    except SystemUser.DoesNotExist:
        return redirect("login")

    with connection.cursor() as cursor:
        cursor.execute("SET app.current_user_id = %s", [user.id])
        cursor.execute("SET app.current_user_role = %s", [user.role])
        if user.role == "admin":
            cursor.execute("SET ROLE admin_role")
        else:
            cursor.execute("SET ROLE doctor_role")

    available_tables = list(TABLES.keys())
    if user.role == "doctor":
        for forbidden in ("patients", "aliases",
                          "action_logs", "encryption_keys"):
            if forbidden in available_tables:
                available_tables.remove(forbidden)

    selected_table = request.GET.get("table")
    if selected_table not in available_tables:
        selected_table = None

    table_data = None
    columns = None

    if selected_table:
        model = TABLES[selected_table]

        fields = []
        for field in model._meta.fields:
            if field.get_internal_type() == "ForeignKey":
                fields.append(field.attname)
            else:
                fields.append(field.name)

        columns = fields
        if selected_table == "doctors":
            table_data = model.objects.select_related('user').values(*fields, 'user_id')
        else:
            table_data = model.objects.values(*fields)

    add_admin_form = AddAdminForm()
    add_doctor_form = AddDoctorForm()

    return render(request, "dashboard.html", {
        "tables": available_tables,
        "selected_table": selected_table,
        "table_data": table_data,
        "columns": columns,
        "current_user": user,
        "form_admin": add_admin_form,
        "form_doctor": add_doctor_form,
        "ADD_ALLOWED_TABLES": ADD_ALLOWED_TABLES
    })


@login_required
def add_employee(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login")
    try:
        user = SystemUser.objects.get(id=user_id)
    except SystemUser.DoesNotExist:
        return redirect("login")

    if user.role != "admin":
        return redirect("dashboard")

    role = request.POST.get("role") or request.GET.get("role") or "admin"

    if request.method == "POST":
        if role == "admin":
            form = AddAdminForm(request.POST)
        else:
            form = AddDoctorForm(request.POST)

        if form.is_valid():
            data = form.cleaned_data
            new_user = SystemUser.objects.create(
                full_name=data["full_name"],
                email=data["email"],
                hashed_password=data["hashed_password"],
                role=role
            )
            if role == "doctor":
                parts = data["full_name"].split(" ", 1)
                first_name = parts[0]
                last_name = parts[1] if len(parts) > 1 else ""
                Doctor.objects.create(
                    user=new_user,
                    first_name=first_name,
                    last_name=last_name,
                    specialization=data["specialization"],
                    license_number=data["license_number"],
                    phone=data["phone"],
                    email=new_user.email
                )
            return redirect("dashboard")
    else:
        if role == "admin":
            form = AddAdminForm()
        else:
            form = AddDoctorForm()

    return render(request, "add_employee.html", {
        "form": form,
        "role": role
    })


from openpyxl.utils import get_column_letter


@login_required
def export_excel(request):
    user = SystemUser.objects.get(id=request.session["user_id"])
    if user.role != "admin":
        return redirect("dashboard")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    TABLES_TO_EXPORT = {k: v for k, v in TABLES.items() if k != "encryption_keys"}

    for table_name, model in TABLES_TO_EXPORT.items():
        ws = wb.create_sheet(title=table_name)
        fields = [f.name for f in model._meta.fields]

        ws.append(fields)

        for obj in model.objects.all():
            row = []
            for f in fields:
                val = getattr(obj, f)
                if hasattr(val, "__str__"):
                    val = str(val)
                row.append(val)
            ws.append(row)

        # автоширина колонок
        for i, column in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2  # +2 для отступа

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
    return response


@login_required
def edit_row(request, table, row_id):
    model = TABLES.get(table)
    if not model:
        return redirect("dashboard")

    user = SystemUser.objects.get(id=request.session["user_id"])

    if user.role != "admin":
        return redirect("dashboard")

    if table == "doctors":
        row = get_object_or_404(model, user_id=row_id)
    else:
        row = get_object_or_404(model, id=row_id)

    doctors = Doctor.objects.all()
    aliases = Alias.objects.all()
    visits = Visit.objects.all()

    if request.method == "POST":
        for field in model._meta.fields:
            name = field.name

            if name in ("id", "hashed_password", "created_at", "updated_at"):
                continue

            value = request.POST.get(name)
            print(value)

            if name == "doctor" and value:
                row.doctor_id = int(value)
            elif name == "alias" and value:
                row.alias_id = int(value)
            elif name == "visit" and value:
                row.visit_id = int(value)
            elif value not in ("", None):
                setattr(row, name, value)

        row.save()
        return redirect("dashboard")

    fields = {}
    role_choices = None
    status_choices = None

    for field in model._meta.fields:
        if field.name not in ("id", "hashed_password", "created_at", "updated_at"):
            fields[field.name] = getattr(row, field.name)

        if field.name == "role":
            role_choices = field.choices

        elif field.name == "status":
            status_choices = field.choices

    return render(request, "edit_row.html",
                  {"fields": fields,
                   "table": table,
                   "row_id": row_id,
                   "role_choices": role_choices,
                   "status_choices": status_choices,
                   "doctors": doctors,
                   "aliases": aliases,
                   "visits": visits, })


@login_required
def delete_row(request):
    user = SystemUser.objects.get(id=request.session["user_id"])
    if user.role != "admin":
        return redirect("dashboard")

    if request.method == "POST":
        table = request.POST.get("table")
        row_id = request.POST.get("row_id")
        model = TABLES.get(table)
        if model:
            obj = get_object_or_404(model, id=row_id)
            obj.delete()

    return redirect("dashboard")


@login_required
def add_row(request, table):
    user = SystemUser.objects.get(id=request.session["user_id"])

    if user.role != "admin":
        return redirect("dashboard")

    if table not in ADD_ALLOWED_TABLES:
        return redirect("dashboard")

    model = TABLES.get(table)
    if not model:
        return redirect("dashboard")

    doctors = Doctor.objects.all()
    aliases = Alias.objects.all()
    visits = Visit.objects.all()

    if request.method == "POST":
        obj = model()

        for field in model._meta.fields:
            name = field.name

            if name in ("id", "created_at", "updated_at", "hashed_password"):
                continue

            value = request.POST.get(name)
            if not value:
                continue

            # ForeignKey — сохраняем через *_id
            if field.is_relation and field.many_to_one:
                setattr(obj, f"{name}_id", value)
            else:
                setattr(obj, name, value)

        obj.save()
        return redirect(f"/dashboard/?table={table}")

    fields = []
    status_choices = None

    for field in model._meta.fields:
        if field.name in ("id", "created_at", "updated_at"):
            continue

        fields.append(field.name)

        if field.name == "status":
            status_choices = field.choices

    return render(request, "add_row.html", {
        "table": table,
        "fields": fields,
        "status_choices": status_choices,
        "doctors": doctors,
        "aliases": aliases,
        "visits": visits,
    })
